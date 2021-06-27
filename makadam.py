import json
import os
from time import time

import click
import cv2
import easyocr
import joblib
import numpy as np
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from utils import (
    apply_overlay,
    create_video_capture,
    image_preprocessing,
    returnCameraIndexes,
)

RESOLUTION = (1280, 720)
RESULTS_TABLE_TL_CORNER = (675, 50)  # top left corner of results table at the end of the race
RESULTS_TABLE_BR_CORNER = (945, 670)  # bottom right corner of results table at the end of the race
POSITION_TL_CORNER = (1078, 574)  # top left corner of position indicator
POSITION_BR_CORNER = (1195, 676)  # bottom right corner of position indicator
DEFAULT_WB_NAME = "default_workbook.xlsx"
DETECTION_THRESHOLD = 0.6


def validate_video_port(value):
    ok_ports = returnCameraIndexes()

    if value == -1:  # default value, prompt the user
        click.secho("The following video ports are active :", fg="green")

        for i in ok_ports:
            if i == 0:
                click.echo(
                    "- Port 0 (if you are using a laptop, this is probably the integrated camera)"
                )
            else:
                click.echo(f"- Port {i}")

        click.secho("Please choose a port", fg="blue")
        value = int(click.prompt("Port"))

    if value not in ok_ports:
        raise click.BadParameter("port must be available")
    else:
        return value


def show_tutorial():
    click.echo("\n")
    click.secho("How to use :", bg="white", fg="black")
    click.echo(
        "Press "
        + click.style("SPACE", bold=True)
        + " to "
        + click.style("start", fg="green")
        + " recording your position"
    )
    click.echo(
        "Press "
        + click.style("SPACE", bold=True)
        + " again to "
        + click.style("stop", fg="red")
        + " recording and save the records to a file"
    )
    click.echo(
        "Press "
        + click.style("Q", bold=True)
        + " to "
        + click.style("exit", fg="magenta")
        + " the program"
    )
    click.echo("")


# ----------------------------
# Define a group of commands. Allows to have multiple commands after `makadam`, like `makadam live` or `makadam scores`
@click.group()
def cli():
    pass


# ----------------------------
@cli.command()
@click.argument("img-dir", type=click.Path(exists=True))
@click.argument("players", nargs=-1, required=True)
@click.option(
    "-w",
    "--workbook",
    default=DEFAULT_WB_NAME,
    show_default=True,
    type=click.Path(),
    help="Path to the workbook to fill. If it does not exist, it will be created. Must include the extension",
)
@click.option(
    "--save-copy/--no-save-copy",
    default=False,
    show_default=True,
    help="Save the new workbook as a copy or overwrite the old one",
)
@click.option(
    "--header/--no-header",
    default=True,
    show_default=True,
    help="Add a header row with the player names",
)
def scores(img_dir, players, workbook, save_copy, header):
    """Read the finish grid from MarioKart screenshots and fill an Excel workbook with the results

    IMG-DIR is the path to the directory where your screenshots are saved

    PLAYERS is the names of the players to write in the Excel workbook, separated by spaces
    """

    # --------- Load the Workbook or create a new one -----------
    try:
        wb = load_workbook(workbook)
    except FileNotFoundError:
        wb = Workbook()
        click.echo(f"[INFO] The file named '{workbook}' has not been found, so it will be created")

    # TODO : read the data already present in the workbook
    # TODO : append the new data to the correct place (existing columns that match with the player)
    ws = wb.active
    if header:
        ws.append(players)  # Add headers

    # --------- Verify the validity of img_dir ----------
    try:
        img_list = os.listdir(img_dir)
    except NotADirectoryError:
        click.echo()
        raise click.BadParameter(message=f"'{img_dir}' is not a directory.")

    # --------- Create the OCR reader ----------
    reader = easyocr.Reader(["fr"])

    # -------- Iterate over all files in img_dir ---------
    for img_name in tqdm(img_list, desc="Analysing images"):
        if img_name.endswith(".jpg"):
            img = cv2.imread(os.path.join(img_dir, img_name), -1)
        else:  # If file is not an image, skip the iteration
            continue

        # ------- Preprocess and read the image ---------
        img = image_preprocessing(img, RESULTS_TABLE_TL_CORNER, RESULTS_TABLE_BR_CORNER)
        results = reader.readtext(
            img,
            allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789",
            detail=1,
        )
        finish_grid = [item[1] for item in results]

        # ------- for Each player wanted by the user, assign a position -------
        positions = []
        for player in players:
            try:
                positions.append(finish_grid.index(player) + 1)
            except ValueError:  # happens when the OCR did not succeed
                click.echo(
                    f"\n\n[WARN] Player '{player}' was not found in the finish grid :( The image I had trouble with is {os.path.abspath(img_name)}\n"
                )
                positions.append("NA")
        # add the positions to a new row in the worksheet
        ws.append(positions)

    # ------- Save the workbook ---------
    if save_copy:
        wb.save(workbook[:-5] + "_copy.xlsx")
    else:
        wb.save(workbook)


@cli.command()
@click.option("--video-port", type=click.INT, default=-1)
def live(video_port):
    """Extract position data from live MarioKart video. Results are dumped in a file."""
    video_port = validate_video_port(video_port)

    video = create_video_capture(video_port, RESOLUTION)

    with open("trained_model.pkl", "rb") as f:
        model = joblib.load(f)

    show_tutorial()

    is_computing = False
    last_10_pos = []
    while True:
        ret, frame = video.read()
        if ret == 0:
            print("video finished or error")
            break

        # Display a red circle when not doing anything and a green circle when computing
        cv2.circle(
            frame,  # image
            (RESOLUTION[0] // 2, RESOLUTION[1] // 10),  # center
            40,  # radius
            (0, 255, 0) if is_computing else (0, 0, 255),  # color
            -1,  # thickness (-1 means filled)
        )

        if is_computing:
            # crop the image to focus on the position
            cropped_frame = frame[
                POSITION_TL_CORNER[1] : POSITION_BR_CORNER[1],
                POSITION_TL_CORNER[0] : POSITION_BR_CORNER[0],
            ]

            gray_cropped_frame = cv2.cvtColor(cropped_frame, cv2.COLOR_BGR2GRAY)

            # Get the probabilities of each category
            # looks like : [5%, 5%, 60%, 5%, 0%, 3%, 10%, 5%, 5%, 0%, 0%, 2%]
            probabilities = list(model.predict_proba(np.array([gray_cropped_frame]))[0])

            # The best prediction's index correspond to the position in the race
            best_prediction = max(probabilities)
            position = probabilities.index(best_prediction) + 1

            # Define a threshold to accept a prediction or not
            if best_prediction > DETECTION_THRESHOLD:

                # Save the last 5 position to smooth the recordings
                last_10_pos.append(position)
                if len(last_10_pos) > 10:
                    last_10_pos.pop(0)

                print(last_10_pos)

                if all([x == position for x in last_10_pos]):
                    records[time() - start_time] = position

                # simple display on the video
                cv2.putText(
                    frame,
                    str(position),
                    POSITION_TL_CORNER,
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,
                    (255, 255, 255),
                    3,
                )
            else:

                cv2.putText(
                    frame,
                    "???",
                    POSITION_TL_CORNER,
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,
                    (255, 255, 255),
                    3,
                )

        cv2.imshow("live", frame)

        # wait 1 milisecond for a key press
        key = cv2.waitKey(1)
        # Press 'q' to quit.
        if key == ord("q"):
            break
        # Press 'Space' to toggle the position computing
        if key == ord(" "):
            # Toggle from recording to not recording
            if is_computing == True:
                # save results
                with open("records.json", "w+") as f:
                    json.dump(records, f)
                is_computing = False  # toggle
            # toggle from not recording to recording
            elif is_computing == False:
                records = {}  # reset the records
                start_time = time()  # reset the time
                is_computing = True  # toggle

    video.release()
    cv2.destroyAllWindows()


if __name__ == "__main__":
    scores()
